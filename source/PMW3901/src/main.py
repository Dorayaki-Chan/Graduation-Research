#!/usr/bin/env python
import time
import argparse
from pmw3901 import PMW3901, BG_CS_FRONT_BCM, BG_CS_BACK_BCM

from openpyxl import Workbook
import os

# ワークブックを作成
workbook = Workbook()

# アクティブなシートを取得
sheet = workbook.active

# シート名を設定
sheet.title = "Sheet1"

# parser = argparse.ArgumentParser()
# parser.add_argument('--board', type=str,
#                     choices=['pmw3901', 'paa5100'],
#                     required=True,
#                     help='Breakout type.')
# parser.add_argument('--rotation', type=int,
#                     default=0, choices=[0, 90, 180, 270],
#                     help='Rotation of sensor in degrees.')
# parser.add_argument('--spi-slot', type=str,
#                     default='front', choices=['front', 'back'],
#                     help='Breakout Garden SPI slot.')
# args = parser.parse_args()

# Pick the right class for the specified breakout
SensorClass = PMW3901 # if args.board == 'pmw3901' else PAA5100

flo = SensorClass(spi_port=0, spi_cs_gpio=BG_CS_FRONT_BCM)
flo.set_rotation(0)     # Rotation of sensor in degrees
                        # choices=[0, 90, 180, 270]

tx = 0
ty = 0

lists = []

try:
    while True:
        try:
            x, y = flo.get_motion()
        except RuntimeError:
            continue
        tx += x
        ty += y
        print("Relative: x {:03d} y {:03d} | Absolute: x {:03d} y {:03d}".format(x, y, tx, ty))
        lists.append([x, y, tx, ty])
        time.sleep(0.01)
except KeyboardInterrupt:
    # logフォルダーのパス
    log_folder_path = "./log"
    # logフォルダーが存在しない場合は作成する
    if not os.path.exists(log_folder_path):
        os.makedirs(log_folder_path)
    
    # セルに値を設定
    for i, list in enumerate(lists):
        sheet.cell(row=i+1, column=1).value = list[0]
        sheet.cell(row=i+1, column=2).value = list[1]
        sheet.cell(row=i+1, column=3).value = list[2]
        sheet.cell(row=i+1, column=4).value = list[3]

    # Excelファイルを保存
    file_path = os.path.join(log_folder_path, "coordinates.xlsx")
    workbook.save(file_path)
    pass
