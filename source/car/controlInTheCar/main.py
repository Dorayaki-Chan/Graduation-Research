#!/usr/bin/env python
import time
import argparse
from pmw3901 import PMW3901, BG_CS_FRONT_BCM, BG_CS_BACK_BCM

from openpyxl import Workbook
import os

import math

OPTICAL_KEISUU = 0.188679245

# ワークブックを作成
workbook = Workbook()

# アクティブなシートを取得
sheet = workbook.active

# シート名を設定
sheet.title = "Sheet1"

# Pick the right class for the specified breakout
SensorClass = PMW3901 

flo = SensorClass(spi_port=0, spi_cs_gpio=BG_CS_FRONT_BCM)
flo.set_rotation(0)     # Rotation of sensor in degrees
                        # choices=[0, 90, 180, 270]

tx = 0
ty = 0

lists = []

# 経過時間を計測するために開始時刻を保存
start = time.time()

# 経過時間を秒で返す
def get_elapsed_time():
    return time.time() - start

try:
    pass
except KeyboardInterrupt:
    # logフォルダーのパス
    log_folder_path = "./log"
    # logフォルダーが存在しない場合は作成する
    if not os.path.exists(log_folder_path):
        os.makedirs(log_folder_path)
    
    # セルに値を設定
    for i, list in enumerate(lists):
        sheet.cell(row=i+1, column=1).value = list[0]
        sheet.cell(row=i+1, column=2).value = list[1]
        sheet.cell(row=i+1, column=3).value = list[2]
        sheet.cell(row=i+1, column=4).value = list[3]

    # Excelファイルを保存
    file_path = os.path.join(log_folder_path, "coordinates.xlsx")
    workbook.save(file_path)
    pass

# ロボット自身の動きを制御するクラス
class DriveTheCar:
    def __init__(self):
        # ロボットの座標/角度を初期化
        self.tx = 0
        self.ty = 0
        self.angle = 0

        self.x = 0
        self.y = 0
        
        self.speed = 0

        # 経過時間を計測するために開始時刻を保存
        self.start = time.time()
    
    # 経過時間を秒で返す
    def __get_elapsed_time(self):
        return time.time() - self.start
    
    def main_loop(self):
        while True:
            try:
                x, y = flo.get_motion()
            except RuntimeError:
                continue
            self.x += x * OPTICAL_KEISUU
            self.y += y * OPTICAL_KEISUU
            # 経過時間を表示
            elapsed_time = time.time() - self.start

            angle = (abs(self.x) * 360) / (2 * math.pi * 45)
            print("Elapsed_time:{0} [sec] | Absolute mm: x {:03d} y {:03d} | Absolute: x {:03d} y {:03d} | Rotation: {0}".format(elapsed_time, self.x, self.y, tx, ty, angle))
            lists.append([get_elapsed_time(), x, y, tx, ty])
            time.sleep(0.01)    
    
    # ロボットの座標/角度を更新する
    def __update(self, x, y, angle):
        self.tx += x * OPTICAL_KEISUU
        self.ty += y * OPTICAL_KEISUU
        self.angle += angle

    def __how_rotation(self, x, y):
        dif_x = x - self.tx
        dif_y = y - self.ty
        
        return math.degrees(math.atan2(dif_y, dif_x))